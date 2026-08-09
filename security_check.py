"""
Checks prospect websites for basic security / trustworthiness red flags
before they're run through the outreach automation (find_contact_links.py,
fill_contact_form.py), and writes a color-coded Excel report.

Checks performed (no paid API required, except where noted):
    - Reachability (does the site respond at all, over HTTPS or HTTP)
    - SSL certificate validity, expiry, and whether HTTPS works at all
    - Domain age via WHOIS (very recently registered domains are riskier)
    - Parked / thin-content pages (placeholder pages, "domain for sale", etc.)
    - Redirects to an unrelated domain
    - Google Safe Browsing blocklist check (OPTIONAL — only runs if you set
      the GOOGLE_SAFE_BROWSING_API_KEY environment variable to a free key
      from https://developers.google.com/safe-browsing/v4/get-started;
      never hardcode the key into this file)

This runs in "aggressive" mode: any single flag is enough to mark a site
Medium or High risk. High-severity flags (unreachable, invalid certificate,
Safe Browsing hit, parked domain) mark a site High risk (colored red in the
report); anything else that's flagged is Medium risk (colored amber); sites
with no flags are Clean (colored green).

Input:
    Same tracker workbook as find_contact_links.py (default:
    "Outreach tracker.xlsx"), auto-detecting company/website columns.

Output:
    "security_report_<YYYYMMDD_HHMMSS>.xlsx", sorted High risk first.

Usage:
    python security_check.py
"""

import os
import socket
import ssl
import time
from datetime import datetime, timezone
from urllib.parse import urlparse

import pandas as pd
import requests
import urllib3
from bs4 import BeautifulSoup
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    import whois as whois_lookup  # python-whois
except ImportError:
    whois_lookup = None

from find_contact_links import load_records  # reuse the same tracker-reading logic

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# ---------------------------------------------------------------------------
# CONFIG
# ---------------------------------------------------------------------------
INPUT_FILE = "Outreach tracker.xlsx"
OUTPUT_FILE_PREFIX = "security_report"

REQUEST_TIMEOUT_SECONDS = 10
DELAY_BETWEEN_REQUESTS_SECONDS = 1.0

USER_AGENT = (
    "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
    "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
)
HEADERS = {"User-Agent": USER_AGENT}

# Set the GOOGLE_SAFE_BROWSING_API_KEY environment variable to enable this
# check. Left unset, the Safe Browsing column just reports "not_checked".
GOOGLE_SAFE_BROWSING_API_KEY = os.environ.get("GOOGLE_SAFE_BROWSING_API_KEY", "")
SAFE_BROWSING_ENDPOINT = "https://safebrowsing.googleapis.com/v4/threatMatches:find"

NEW_DOMAIN_THRESHOLD_DAYS = 180
CERT_EXPIRING_SOON_DAYS = 30
THIN_CONTENT_MIN_CHARS = 200

PARKED_DOMAIN_PHRASES = [
    "domain is for sale", "buy this domain", "this domain is parked",
    "future home of", "under construction", "coming soon",
    "domain may be for sale", "this web page is parked",
    "the domain has expired", "related searches",
]

# Any flag in this set makes a site High risk; every other flag is Medium.
HIGH_SEVERITY_FLAG_CODES = {"unreachable", "invalid_certificate", "safe_browsing_hit", "parked_domain"}

FILL_HIGH = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
FILL_MEDIUM = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
FILL_CLEAN = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
FONT_HIGH = Font(color="9C0006")
FONT_MEDIUM = Font(color="9C6500")
FONT_CLEAN = Font(color="006100")
# ---------------------------------------------------------------------------


def _same_registrable_domain(host_a: str, host_b: str) -> bool:
    """Naive registrable-domain comparison (last two dot-separated labels,
    stripping a leading www.). Doesn't handle multi-part TLDs like .co.uk,
    but is dependency-free and good enough to catch a redirect to an
    unrelated domain."""
    def base(h):
        h = (h or "").lower()
        if h.startswith("www."):
            h = h[4:]
        parts = h.split(".")
        return ".".join(parts[-2:]) if len(parts) >= 2 else h
    return base(host_a) == base(host_b)


def fetch_page(url: str, session: requests.Session) -> dict:
    """Try HTTPS, falling back to HTTP if HTTPS fails outright, and to an
    unverified HTTPS request if only certificate validation fails (so we can
    still inspect the page while flagging the bad cert separately)."""
    result = {"reachable": False, "https_ok": False, "final_url": url, "html": "", "error": ""}
    try:
        resp = session.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT_SECONDS, allow_redirects=True)
        result.update(reachable=True, https_ok=url.startswith("https://"), final_url=resp.url, html=resp.text)
        return result
    except requests.exceptions.SSLError:
        try:
            resp = session.get(url, headers=HEADERS, timeout=REQUEST_TIMEOUT_SECONDS, allow_redirects=True, verify=False)
            result.update(reachable=True, https_ok=False, final_url=resp.url, html=resp.text)
            return result
        except requests.exceptions.RequestException as exc2:
            result["error"] = str(exc2)
            return result
    except requests.exceptions.RequestException as exc:
        if url.startswith("https://"):
            http_url = "http://" + url[len("https://"):]
            try:
                resp = session.get(http_url, headers=HEADERS, timeout=REQUEST_TIMEOUT_SECONDS, allow_redirects=True)
                result.update(reachable=True, https_ok=False, final_url=resp.url, html=resp.text)
                return result
            except requests.exceptions.RequestException as exc2:
                result["error"] = str(exc2)
                return result
        result["error"] = str(exc)
        return result


def _parse_notafter(not_after: str) -> datetime:
    return datetime.strptime(not_after, "%b %d %H:%M:%S %Y %Z").replace(tzinfo=timezone.utc)


def inspect_certificate(hostname: str, port: int = 443, timeout: int = REQUEST_TIMEOUT_SECONDS):
    """Returns (status, days_until_expiry). status is one of:
    valid, expiring_soon, expired, invalid, no_https, error."""
    context = ssl.create_default_context()
    try:
        with socket.create_connection((hostname, port), timeout=timeout) as sock:
            with context.wrap_socket(sock, server_hostname=hostname) as ssock:
                cert = ssock.getpeercert()
        not_after = _parse_notafter(cert["notAfter"])
        days_left = (not_after - datetime.now(timezone.utc)).days
        if days_left < 0:
            return "expired", days_left
        if days_left <= CERT_EXPIRING_SOON_DAYS:
            return "expiring_soon", days_left
        return "valid", days_left
    except ssl.SSLCertVerificationError:
        return "invalid", None
    except (socket.timeout, socket.gaierror, ConnectionRefusedError, OSError):
        return "no_https", None
    except Exception:
        return "error", None


def domain_age_days(hostname: str):
    """Returns (age_days_or_None, error_message_or_None)."""
    if whois_lookup is None:
        return None, "python-whois not installed"
    try:
        record = whois_lookup.whois(hostname)
        created = record.creation_date
        if isinstance(created, list):
            created = created[0]
        if created is None:
            return None, "unavailable"
        if created.tzinfo is None:
            created = created.replace(tzinfo=timezone.utc)
        return (datetime.now(timezone.utc) - created).days, None
    except Exception as exc:
        return None, str(exc)


def check_safe_browsing(url: str, session: requests.Session):
    """Returns (status, detail). status: not_checked, clean, flagged, error."""
    if not GOOGLE_SAFE_BROWSING_API_KEY:
        return "not_checked", None
    payload = {
        "client": {"clientId": "invocursor-outreach-security-check", "clientVersion": "1.0"},
        "threatInfo": {
            "threatTypes": ["MALWARE", "SOCIAL_ENGINEERING", "UNWANTED_SOFTWARE", "POTENTIALLY_HARMFUL_APPLICATION"],
            "platformTypes": ["ANY_PLATFORM"],
            "threatEntryTypes": ["URL"],
            "threatEntries": [{"url": url}],
        },
    }
    try:
        resp = session.post(
            SAFE_BROWSING_ENDPOINT,
            params={"key": GOOGLE_SAFE_BROWSING_API_KEY},
            json=payload,
            timeout=REQUEST_TIMEOUT_SECONDS,
        )
        resp.raise_for_status()
        matches = resp.json().get("matches", [])
        if matches:
            threat_types = ", ".join(sorted({m.get("threatType", "UNKNOWN") for m in matches}))
            return "flagged", threat_types
        return "clean", None
    except requests.exceptions.RequestException as exc:
        return "error", str(exc)


def check_website_security(company: str, url: str, session: requests.Session) -> dict:
    flags = []
    hostname = urlparse(url).hostname or url
    fetch = fetch_page(url, session)
    cert_status, cert_days = inspect_certificate(hostname)
    content_length = 0

    if not fetch["reachable"]:
        flags.append(("unreachable", f"Site did not respond ({fetch['error']})"))
    else:
        if not fetch["https_ok"]:
            flags.append(("no_https", "Site does not serve a working, valid HTTPS connection"))

        final_host = urlparse(fetch["final_url"]).hostname or ""
        if final_host and not _same_registrable_domain(hostname, final_host):
            flags.append(("suspicious_redirect", f"Redirects to a different domain ({final_host})"))

        text = BeautifulSoup(fetch["html"], "html.parser").get_text(" ", strip=True) if fetch["html"] else ""
        content_length = len(text)
        if content_length < THIN_CONTENT_MIN_CHARS:
            flags.append(("thin_content", f"Page has very little visible content ({content_length} chars)"))
        lowered = text.lower()
        if any(phrase in lowered for phrase in PARKED_DOMAIN_PHRASES):
            flags.append(("parked_domain", "Page text matches common parked-domain/placeholder wording"))

    if cert_status == "expired":
        flags.append(("invalid_certificate", f"SSL certificate expired {abs(cert_days)} days ago"))
    elif cert_status == "invalid":
        flags.append(("invalid_certificate", "SSL certificate failed validation"))
    elif cert_status == "expiring_soon":
        flags.append(("cert_expiring_soon", f"SSL certificate expires in {cert_days} days"))

    age_days, _age_error = domain_age_days(hostname)
    if age_days is not None and age_days < NEW_DOMAIN_THRESHOLD_DAYS:
        flags.append(("new_domain", f"Domain registered only {age_days} days ago"))

    sb_status, sb_detail = check_safe_browsing(fetch["final_url"] or url, session)
    if sb_status == "flagged":
        flags.append(("safe_browsing_hit", f"Listed on Google Safe Browsing ({sb_detail})"))

    if any(code in HIGH_SEVERITY_FLAG_CODES for code, _ in flags):
        risk_level = "High"
    elif flags:
        risk_level = "Medium"
    else:
        risk_level = "Clean"

    return {
        "Company": company,
        "Website": url,
        "Risk Level": risk_level,
        "Flags": "; ".join(desc for _code, desc in flags),
        "Reachable": "Yes" if fetch["reachable"] else "No",
        "HTTPS OK": "Yes" if fetch["https_ok"] else "No",
        "Certificate Status": cert_status,
        "Cert Days Remaining": cert_days if cert_days is not None else "",
        "Domain Age (days)": age_days if age_days is not None else "",
        "Content Length (chars)": content_length,
        "Final URL": fetch["final_url"],
        "Safe Browsing": sb_status,
        "Error": fetch["error"],
    }


def check_websites(records: list, delay: float = DELAY_BETWEEN_REQUESTS_SECONDS) -> list:
    session = requests.Session()
    rows = []
    for i, (company, website) in enumerate(records, start=1):
        print(f"[{i}/{len(records)}] Checking {company or website} -> {website}")
        rows.append(check_website_security(company, website, session))
        if i < len(records):
            time.sleep(delay)
    return rows


_RISK_SORT_ORDER = {"High": 0, "Medium": 1, "Clean": 2}


def build_report(rows: list) -> pd.DataFrame:
    columns = [
        "Company", "Website", "Risk Level", "Flags", "Reachable", "HTTPS OK",
        "Certificate Status", "Cert Days Remaining", "Domain Age (days)",
        "Content Length (chars)", "Final URL", "Safe Browsing", "Error",
    ]
    ordered = sorted(rows, key=lambda r: _RISK_SORT_ORDER[r["Risk Level"]])
    return pd.DataFrame(ordered, columns=columns)


def write_report(df: pd.DataFrame, output_file: str) -> None:
    with pd.ExcelWriter(output_file, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Security Report")
        ws = writer.sheets["Security Report"]

        header_fill = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
        for col_idx, col_name in enumerate(df.columns, start=1):
            cell = ws.cell(row=1, column=col_idx)
            cell.font = Font(bold=True, color="FFFFFF")
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            ws.column_dimensions[get_column_letter(col_idx)].width = max(14, min(40, len(str(col_name)) + 4))
        ws.freeze_panes = "A2"

        risk_col = list(df.columns).index("Risk Level") + 1
        for row_idx in range(2, ws.max_row + 1):
            risk = ws.cell(row=row_idx, column=risk_col).value
            fill, font = {
                "High": (FILL_HIGH, FONT_HIGH),
                "Medium": (FILL_MEDIUM, FONT_MEDIUM),
                "Clean": (FILL_CLEAN, FONT_CLEAN),
            }.get(risk, (None, None))
            if fill is None:
                continue
            for col_idx in range(1, len(df.columns) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = fill
                cell.font = font


def main():
    records = load_records(INPUT_FILE)
    if not records:
        raise ValueError(f"No usable website URLs found in '{INPUT_FILE}'.")

    if not GOOGLE_SAFE_BROWSING_API_KEY:
        print("Note: GOOGLE_SAFE_BROWSING_API_KEY is not set — Safe Browsing checks will be skipped.\n")

    rows = check_websites(records)
    df = build_report(rows)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = f"{OUTPUT_FILE_PREFIX}_{timestamp}.xlsx"
    write_report(df, output_file)

    high = sum(1 for r in rows if r["Risk Level"] == "High")
    medium = sum(1 for r in rows if r["Risk Level"] == "Medium")
    clean = sum(1 for r in rows if r["Risk Level"] == "Clean")
    print(f"\nWrote {len(df)} rows to {output_file} ({high} High, {medium} Medium, {clean} Clean)")


if __name__ == "__main__":
    main()
